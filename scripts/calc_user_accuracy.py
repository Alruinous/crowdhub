#!/usr/bin/env python3
"""Compute per-user daily accuracy from a wide xlsx table.

Input format example:
- First column: name
- For each day there are two columns:
  - 01-26_total
  - 01-26_correct

Output:
- A new xlsx file under test/accuracy by default
- Sheet 1: per-user daily accuracy
- Sheet 2: daily aggregate accuracy across all users

Accuracy is computed as correct / total.
If total is 0 or missing, the output cell is left blank.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError as exc:  # pragma: no cover - dependency guidance
    raise SystemExit(
        "缺少 openpyxl，请先安装：pip install openpyxl"
    ) from exc


HEADER_RE = re.compile(r"^(?P<day>\d{2}-\d{2})_(?P<kind>total|correct)$")
PROJECT_ROOT = Path(__file__).resolve().parents[1]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate per-user daily accuracy from a wide xlsx file."
    )
    parser.add_argument("input", help="Input xlsx file path")
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name or index (default: first sheet)",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="Output xlsx path (default: test/accuracy/<input>_accuracy.xlsx)",
    )
    return parser.parse_args()


def resolve_input_path(raw: str) -> Path:
    path = Path(raw).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"输入文件不存在: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"输入文件必须是 xlsx: {path}")
    return path


def resolve_output_path(input_path: Path, out_arg: Optional[str]) -> Path:
    if out_arg:
        out_path = Path(out_arg).expanduser().resolve()
    else:
        out_dir = PROJECT_ROOT / "test" / "accuracy"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{input_path.stem}_accuracy.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    return out_path


def get_sheet(wb, sheet_arg: Optional[str]):
    if sheet_arg is None:
        return wb[wb.sheetnames[0]]

    if sheet_arg.isdigit():
        index = int(sheet_arg)
        if index < 0 or index >= len(wb.sheetnames):
            raise IndexError(f"sheet 索引越界: {index}")
        return wb[wb.sheetnames[index]]

    if sheet_arg not in wb.sheetnames:
        raise KeyError(f"找不到工作表: {sheet_arg}")
    return wb[sheet_arg]


def read_headers(ws) -> Tuple[List[str], Dict[str, int], List[str]]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if not headers:
        raise ValueError("表头为空")

    header_names = ["" if value is None else str(value).strip() for value in headers]
    day_kinds: Dict[str, Dict[str, int]] = {}
    ordered_days: List[str] = []

    for idx, header in enumerate(header_names):
        match = HEADER_RE.match(header)
        if not match:
                        continue
        day = match.group("day")
        kind = match.group("kind")
        if day not in day_kinds:
            day_kinds[day] = {}
            ordered_days.append(day)
        day_kinds[day][kind] = idx

    if not ordered_days:
        raise ValueError("没有找到形如 01-26_total / 01-26_correct 的列")

    header_index = {name: idx for idx, name in enumerate(header_names)}
    return header_names, header_index, ordered_days


def to_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def compute_accuracy(total, correct) -> Optional[float]:
    total_value = to_float(total)
    correct_value = to_float(correct)
    if total_value is None or total_value <= 0:
        return None
    if correct_value is None:
        correct_value = 0.0
    return correct_value / total_value


def build_output_rows(ws, header_index: Dict[str, int], ordered_days: List[str]):
    name_col = header_index.get("name", 0)
    rows: List[List[object]] = []
    summary = {day: {"total": 0.0, "correct": 0.0} for day in ordered_days}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        name = row[name_col] if name_col < len(row) else None
        if name is None or str(name).strip() == "":
            continue

        out_row: List[object] = [name]
        for day in ordered_days:
            total_idx = header_index.get(f"{day}_total")
            correct_idx = header_index.get(f"{day}_correct")
            total = row[total_idx] if total_idx is not None and total_idx < len(row) else None
            correct = row[correct_idx] if correct_idx is not None and correct_idx < len(row) else None
            total_value = to_float(total)
            correct_value = to_float(correct)
            if total_value is not None:
                summary[day]["total"] += total_value
            if correct_value is not None:
                summary[day]["correct"] += correct_value
            out_row.append(compute_accuracy(total, correct))
        rows.append(out_row)

    summary_rows: List[List[object]] = []
    for day in ordered_days:
        total = summary[day]["total"]
        correct = summary[day]["correct"]
        accuracy = correct / total if total > 0 else None
        summary_rows.append([day, total, correct, accuracy])

    return rows, summary_rows


def write_output(
    out_path: Path,
    ordered_days: List[str],
    rows: List[List[object]],
    summary_rows: List[List[object]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "表一_个人正确率"

    header = ["name", *[f"{day}_accuracy" for day in ordered_days]]
    ws.append(header)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    for row in rows:
        ws.append(row)

    for col_idx in range(2, len(header) + 1):
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"

    ws.freeze_panes = "B2"

    summary_ws = wb.create_sheet("表二_每日汇总")
    summary_ws.append(["day", "total", "correct", "accuracy"])
    for cell in summary_ws[1]:
        cell.font = bold_font

    for row in summary_rows:
        summary_ws.append(row)

    for row_idx in range(2, summary_ws.max_row + 1):
        summary_ws.cell(row=row_idx, column=4).number_format = "0.00%"

    summary_ws.freeze_panes = "A2"
    wb.save(out_path)


def main() -> int:
    args = parse_args()
    input_path = resolve_input_path(args.input)
    out_path = resolve_output_path(input_path, args.out)

    wb = load_workbook(input_path, data_only=True)
    ws = get_sheet(wb, args.sheet)

    _, header_index, ordered_days = read_headers(ws)
    rows, summary_rows = build_output_rows(ws, header_index, ordered_days)
    write_output(out_path, ordered_days, rows, summary_rows)

    print(f"已输出：{out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())